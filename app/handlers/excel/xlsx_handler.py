from openpyxl import load_workbook
from typing import Dict, Optional
from ..base.document_handler import DocumentHandler

class XLSXHandler(DocumentHandler):
    """
    Handler for XLSX (Excel 2007 and later) files. Extracts information including 
    the number of sheets in the XLSX workbook.
    """

    def get_info(self, file_path: str, filters: Optional[Dict[str, bool]] = None) -> Optional[Dict[str, str]]:
        """
        Extracts information from an XLSX file, including file details and the number of 
        sheets in the workbook.

        Args:
            file_path (str): The path to the XLSX file.
            filters (Optional[Dict[str, bool]]): Optional dictionary specifying which details to extract.

        Returns:
            Optional[Dict[str, str]]: A dictionary with file information including the 
            number of sheets, or None if an error occurs.
        """
        # Set default filters if not provided
        if filters is None:
            filters = {"sheets": False}

        try:
            file_info = self.extract_file_info(file_path)

            if filters.get('sheets', False):
                workbook = load_workbook(file_path, read_only=True)
                num_sheets = len(workbook.sheetnames)
                file_info.update({"sheets": num_sheets})

            return file_info
        except FileNotFoundError:
            print(f"File not found: {file_path}")
            return None
        except PermissionError:
            print(f"Permission denied: {file_path}")
            return None
        except Exception as e:
            print(f"Error processing {file_path}: {e}")
            return None
